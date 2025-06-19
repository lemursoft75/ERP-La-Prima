import sys
from flask import Flask, render_template, request, redirect, session, jsonify, send_file
import pandas as pd
import os
from datetime import datetime
from openpyxl.workbook import Workbook
import io
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import base64
from io import BytesIO
import threading
import time
import webbrowser



# --- IMPORTS Y CONFIGURACIÓN DE FIREBASE ---
import firebase_admin
from firebase_admin import credentials
from firebase_admin import firestore
# --- FIN IMPORTS Y CONFIGURACIÓN DE FIREBASE ---

app = Flask(__name__, template_folder=os.path.join(os.path.dirname(os.path.abspath(__file__)), 'templates'))
app.secret_key = "supersecreto"


# --- ESTE ES EL BLOQUE DE INICIALIZACIÓN DE FIREBASE QUE DEBE ESTAR UNA SOLA VEZ ---
try:
    # Asegúrate de que 'firebase_credentials.json' esté en la misma carpeta que app.py
    # o especifica la ruta completa.
    cred = credentials.Certificate('firebase_credentials.json')
    firebase_admin.initialize_app(cred)
    db = firestore.client()
    print("Firebase inicializado correctamente.")
except Exception as e:
    print(f"Error al inicializar Firebase: {e}")
    sys.exit(1) # Salir si Firebase falla al inicializar
# --- FIN DEL BLOQUE DE INICIALIZACIÓN DE FIREBASE ---


# Abre directo la app en el Navegador
def open_browser():
    time.sleep(1)
    webbrowser.open_new_tab("http://127.0.0.1:5000")




# --- NUEVAS FUNCIONES PARA INTERACTUAR CON FIRESTORE ---

def load_data_from_firestore(collection_name):
    """
    Carga todos los documentos de una colección de Firestore.
    Retorna una lista de diccionarios, cada diccionario es un documento.
    """
    try:
        docs = db.collection(collection_name).stream()
        data = []
        for doc in docs:
            doc_data = doc.to_dict()
            doc_data['id'] = doc.id # Puedes incluir el ID del documento si lo necesitas
            data.append(doc_data)
        return data
    except Exception as e:
        print(f"Error al cargar datos de la colección '{collection_name}': {e}")
        return []

def save_data_to_firestore(collection_name, document_id, data):
    """
    Guarda o actualiza un documento en una colección de Firestore.
    Si document_id es None, Firestore generará un ID automático.
    """
    try:
        if document_id:
            db.collection(collection_name).document(document_id).set(data)
        else:
            db.collection(collection_name).add(data)
        return True
    except Exception as e:
        print(f"Error al guardar datos en la colección '{collection_name}' con ID '{document_id}': {e}")
        return False

def delete_data_from_firestore(collection_name, document_id):
    """
    Elimina un documento específico de una colección de Firestore.
    """
    try:
        db.collection(collection_name).document(document_id).delete()
        return True
    except Exception as e:
        print(f"Error al eliminar documento '{document_id}' de la colección '{collection_name}': {e}")
        return False



# Recuperacion de datos (esto parece para subidas de archivos, no de datos estructurados)
def get_uploads_folder():
    """ Get absolute path to uploads folder, works for dev and for PyInstaller """
    base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    uploads_path = os.path.join(base_path, "uploads")
    os.makedirs(uploads_path, exist_ok=True)
    return uploads_path


# Carga de datos (productos) - ESTA ES LA FUNCIÓN CLAVE A MODIFICAR
@app.route("/upload-excel", methods=["POST"])
def upload_excel():
    if "file" not in request.files:
        return render_template("register_product.html", message="❌ Error: No se proporcionó un archivo.")

    file = request.files["file"]
    uploads_folder = get_uploads_folder()
    file_path = os.path.join(uploads_folder, file.filename)
    try:
        file.save(file_path)

        df = pd.read_excel(file_path, engine="openpyxl")
        # --- MODIFICACIÓN CLAVE: Cargar productos desde Firestore ---
        existing_products = load_data_from_firestore("products") # Cargar productos existentes de Firestore
        # Convertimos la lista de diccionarios a un conjunto de claves para una búsqueda eficiente
        existing_product_keys = {p.get("clave") for p in existing_products if p.get("clave")}

        products_to_add_count = 0
        for _, row in df.iterrows():
            new_product = {
                "clave": str(row["Clave"]), # Asegúrate de que la clave sea un string si la usas como ID o campo de búsqueda
                "articulo": row["Artículo"],
                "marca": row["Marca"],
                "categoria": row["Categoria"],
                "tamaño": row["Tamaño"],
                "observaciones": row["Observaciones"]
            }
            # Asegurar que no se dupliquen productos por su "clave"
            if new_product["clave"] not in existing_product_keys:
                # --- MODIFICACIÓN CLAVE: Guardar cada producto en Firestore ---
                # Usamos la "clave" del producto como ID del documento en Firestore para fácil referencia.
                # Si la clave no es única, deberías dejar que Firestore genere un ID y almacenar la clave como un campo.
                success = save_data_to_firestore("products", new_product["clave"], new_product)
                if success:
                    products_to_add_count += 1
                    existing_product_keys.add(new_product["clave"]) # Actualizar el conjunto de claves existentes
                else:
                    print(f"Advertencia: No se pudo guardar el producto con clave {new_product['clave']}")

        os.remove(file_path)

        return render_template("register_product.html", message=f"✅ Base actualizada correctamente. Se agregaron {products_to_add_count} productos nuevos.")

    except Exception as e:
        return render_template("register_product.html", message=f"❌ Error procesando archivo: {e}")



# --- NUEVAS FUNCIONES PARA INTERACTUAR CON FIRESTORE (Usuarios) ---

def get_user_from_firestore(username):
    """
    Recupera un usuario de la colección 'users' en Firestore por su nombre de usuario.
    Retorna el diccionario de datos del usuario o None si no existe.
    """
    try:
        user_ref = db.collection('users').document(username)
        user_doc = user_ref.get()
        if user_doc.exists:
            return user_doc.to_dict()
        else:
            return None
    except Exception as e:
        print(f"Error al obtener usuario '{username}' de Firestore: {e}")
        return None

def add_user_to_firestore(username, user_data):
    """
    Agrega un nuevo usuario a la colección 'users' en Firestore.
    Usa el nombre de usuario como ID del documento.
    Retorna True si la operación fue exitosa, False en caso contrario.
    """
    try:
        # Usamos .set() para crear o sobrescribir un documento con un ID específico
        db.collection('users').document(username).set(user_data)
        return True
    except Exception as e:
        print(f"Error al agregar usuario '{username}' a Firestore: {e}")
        return False

def update_user_in_firestore(username, user_data):
    """
    Actualiza los datos de un usuario existente en la colección 'users'.
    Retorna True si la operación fue exitosa, False en caso contrario.
    """
    try:
        # Usamos .set() para sobrescribir el documento con los datos actualizados.
        # Para actualizaciones parciales, podrías usar .update(data) o .set(data, merge=True).
        db.collection('users').document(username).set(user_data)
        return True
    except Exception as e:
        print(f"Error al actualizar usuario '{username}' en Firestore: {e}")
        return False

# --- FIN NUEVAS FUNCIONES PARA FIRESTORE (Usuarios) ---



# ACCESO CON CREDENCIALES

# Página de inicio de sesión
@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]

        # --- CAMBIO CLAVE: Cargar usuario de Firestore ---
        user_data = get_user_from_firestore(username)

        if user_data and user_data.get("password") == password:
            session["user"] = username
            return redirect("/dashboard")
        else:
            return "Credenciales incorrectas, intenta de nuevo"

    return render_template("login.html")


# Página de registro de usuario nuevo
@app.route("/register", methods=["GET", "POST"])
def register():
    message = None

    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]
        question = request.form["question"]
        answer = request.form["answer"]

        # --- CAMBIO CLAVE: Verificar si el usuario ya existe en Firestore ---
        if get_user_from_firestore(username):
            message = "❌ El usuario ya existe. Prueba otro nombre."
        else:
            new_user_data = {
                "password": password,
                "security_question": question,
                "security_answer": answer
            }
            # --- CAMBIO CLAVE: Guardar nuevo usuario en Firestore ---
            if add_user_to_firestore(username, new_user_data):
                return redirect("/")  # Redirige a login después de un registro exitoso
            else:
                message = "❌ Error al registrar el usuario. Intenta de nuevo."  # Mensaje de error si falla el guardado en Firestore

    return render_template("register.html", message=message)


# Página de cambio de contraseña
@app.route("/change-password", methods=["GET", "POST"])
def change_password():
    if request.method == "POST":
        username = request.form["username"]
        question = request.form["question"]
        answer = request.form["answer"]
        new_password = request.form["new_password"]

        # --- CAMBIO CLAVE: Cargar usuario de Firestore ---
        user_data = get_user_from_firestore(username)

        if user_data and \
                user_data.get("security_question") == question and \
                user_data.get("security_answer",
                              "").lower() == answer.lower():  # Usar .get() para evitar KeyError y manejar posible ausencia de campo

            # Actualizar la contraseña en el diccionario de datos del usuario
            user_data["password"] = new_password

            # --- CAMBIO CLAVE: Guardar cambios en Firestore ---
            if update_user_in_firestore(username, user_data):
                return "✅ Contraseña cambiada exitosamente. Ahora puedes iniciar sesión con la nueva contraseña."
            else:
                return "❌ Error al cambiar la contraseña en la base de datos. Intenta de nuevo."
        else:
            return "❌ Datos incorrectos (usuario, pregunta o respuesta de seguridad). No se pudo cambiar la contraseña."

    return render_template("change_password.html")


# MENU PRINCIPAL Y MODULOS

# Página principal del ERP (requiere login)
@app.route("/dashboard")
def dashboard():
    if "user" not in session:
        return redirect("/")
    return render_template("dashboard.html", user=session["user"])



# --- FUNCIONES PARA INTERACTUAR CON FIRESTORE (Productos e Inventario) ---

def get_product_from_firestore(clave):
    """
    Recupera un producto de la colección 'products' por su clave.
    Retorna el diccionario de datos del producto o None si no existe.
    """
    try:
        product_ref = db.collection('products').document(clave)
        product_doc = product_ref.get()
        if product_doc.exists:
            return product_doc.to_dict()
        else:
            return None
    except Exception as e:
        print(f"Error al obtener producto '{clave}' de Firestore: {e}")
        return None

def add_product_to_firestore(clave, product_data):
    """
    Agrega un nuevo producto a la colección 'products'.
    Usa la clave del producto como ID del documento.
    Retorna True si la operación fue exitosa, False en caso contrario.
    """
    try:
        db.collection('products').document(clave).set(product_data)
        return True
    except Exception as e:
        print(f"Error al agregar producto '{clave}' a Firestore: {e}")
        return False

def update_product_in_firestore(clave, product_data):
    """
    Actualiza los datos de un producto existente en la colección 'products'.
    Retorna True si la operación fue exitosa, False en caso contrario.
    """
    try:
        db.collection('products').document(clave).set(product_data) # set() with doc ID overwrites
        return True
    except Exception as e:
        print(f"Error al actualizar producto '{clave}' en Firestore: {e}")
        return False

def delete_product_from_firestore(clave):
    """
    Elimina un producto de la colección 'products'.
    Retorna True si la operación fue exitosa, False en caso contrario.
    """
    try:
        db.collection('products').document(clave).delete()
        return True
    except Exception as e:
        print(f"Error al eliminar producto '{clave}' de Firestore: {e}")
        return False

def get_inventory_item_from_firestore(clave):
    """
    Recupera un ítem de inventario de la colección 'inventory' por su clave.
    Retorna el diccionario de datos del ítem de inventario o None si no existe.
    """
    try:
        inventory_ref = db.collection('inventory').document(clave)
        inventory_doc = inventory_ref.get()
        if inventory_doc.exists:
            return inventory_doc.to_dict()
        else:
            return None
    except Exception as e:
        print(f"Error al obtener ítem de inventario '{clave}' de Firestore: {e}")
        return None

def update_inventory_item_in_firestore(clave, inventory_data):
    """
    Actualiza o crea un ítem de inventario en la colección 'inventory'.
    Retorna True si la operación fue exitosa, False en caso contrario.
    """
    try:
        db.collection('inventory').document(clave).set(inventory_data) # set() will create if not exists, update if exists
        return True
    except Exception as e:
        print(f"Error al actualizar inventario para '{clave}' en Firestore: {e}")
        return False

def delete_inventory_item_from_firestore(clave):
    """
    Elimina un ítem de inventario de la colección 'inventory'.
    Retorna True si la operación fue exitosa, False en caso contrario.
    """
    try:
        db.collection('inventory').document(clave).delete()
        return True
    except Exception as e:
        print(f"Error al eliminar ítem de inventario '{clave}' de Firestore: {e}")
        return False

# --- FIN FUNCIONES PARA FIRESTORE (Productos e Inventario) ---


# PRODUCTOS

# Página de registro de productos
@app.route("/register-product", methods=["GET", "POST"])
def register_product():
    message = None

    if request.method == "POST":
        clave = request.form["clave"]

        # Primero, verifica si el producto ya existe
        if get_product_from_firestore(clave):
            message = "❌ Error: Un producto con esta clave ya existe."
        else:
            new_product = {
                "clave": clave,
                "articulo": request.form["articulo"],
                "marca": request.form["marca"],
                "categoria": request.form["categoria"],
                "tamaño": request.form["tamaño"],
                "observaciones": request.form["observaciones"]
            }

            # Intenta agregar el producto a Firestore
            if add_product_to_firestore(clave, new_product):
                # También inicializa su inventario (existencias y precio)
                initial_inventory = {
                    "clave": clave,
                    "existencias": 0,  # Siempre inicializa en 0 al registrar
                    "precio_unitario": 0  # Siempre inicializa en 0 al registrar
                }
                update_inventory_item_in_firestore(clave, initial_inventory)  # Crea o actualiza el inventario

                message = "✅ Producto registrado correctamente."
            else:
                message = "❌ Error al registrar el producto en la base de datos."

    return render_template("register_product.html", message=message)


# Busqueda de productos
@app.route("/search-product", methods=["POST"])
def search_product():
    clave = request.form["search_clave"]

    # --- CAMBIO CLAVE: Buscar producto en Firestore ---
    product = get_product_from_firestore(clave)

    if product:
        # --- CAMBIO CLAVE: Buscar existencias y precio en inventario de Firestore ---
        inventory_item = get_inventory_item_from_firestore(clave)
        product["existencias"] = inventory_item.get("existencias", 0) if inventory_item else 0
        product["precio_unitario"] = inventory_item.get("precio_unitario", 0) if inventory_item else 0

        return render_template("edit_product.html", product=product, message="✅ Producto encontrado.")

    return render_template("register_product.html", message="❌ Producto no encontrado.")


# Ver productos (usado por AJAX para buscar un producto individualmente)
@app.route("/get-product", methods=["GET"])
def get_product():
    clave = request.args.get("clave")

    # --- CAMBIO CLAVE: Obtener producto de Firestore ---
    product = get_product_from_firestore(clave)

    if product:
        # --- CAMBIO CLAVE: Obtener inventario de Firestore ---
        inventory_item = get_inventory_item_from_firestore(clave)
        existencias = inventory_item.get("existencias", 0) if inventory_item else 0
        precio_unitario = inventory_item.get("precio_unitario", 0) if inventory_item else 0

        return jsonify({
            "found": True,
            "articulo": product["articulo"],
            "marca": product["marca"],
            "categoria": product["categoria"],
            "tamaño": product["tamaño"],
            "observaciones": product.get("observaciones", ""),
            "existencias": existencias,
            "precio_unitario": precio_unitario
        })

    return jsonify({"found": False})


# Pagina de modificacion de productos
@app.route("/edit-product", methods=["POST", "GET"])
def edit_product():
    clave = request.form.get("clave") if request.method == "POST" else request.args.get("clave")

    # --- CAMBIO CLAVE: Obtener producto de Firestore ---
    product = get_product_from_firestore(clave)

    if request.method == "POST" and product:
        print("Datos recibidos:", request.form)  # 🔍 Depuración para ver los datos enviados

        # Actualizar datos del producto
        product["articulo"] = request.form["articulo"]
        product["marca"] = request.form["marca"]
        product["categoria"] = request.form["categoria"]
        product["tamaño"] = request.form["tamaño"]
        product["observaciones"] = request.form["observaciones"]

        # Actualizar existencias y precio unitario del inventario
        existencias_nuevas = int(request.form.get("existencias", 0))
        precio_unitario_nuevo = float(
            request.form.get("precio_unitario", 0))  # Asumiendo que el formulario también tiene este campo

        inventory_data = {
            "clave": clave,
            "existencias": existencias_nuevas,
            "precio_unitario": precio_unitario_nuevo
        }

        # --- CAMBIO CLAVE: Guardar cambios en Firestore ---
        product_updated = update_product_in_firestore(clave, product)
        inventory_updated = update_inventory_item_in_firestore(clave, inventory_data)

        if product_updated and inventory_updated:
            product["existencias"] = existencias_nuevas  # Para mostrar en la plantilla
            product["precio_unitario"] = precio_unitario_nuevo  # Para mostrar en la plantilla
            return render_template("edit_product.html", product=product, message="✅ Cambios guardados correctamente.")
        else:
            return render_template("edit_product.html", product=product,
                                   message="❌ Error al guardar los cambios en la base de datos.")

    if product:
        # --- CAMBIO CLAVE: Cargar existencias y precio para mostrar en el formulario GET ---
        inventory_item = get_inventory_item_from_firestore(clave)
        product["existencias"] = inventory_item.get("existencias", 0) if inventory_item else 0
        product["precio_unitario"] = inventory_item.get("precio_unitario", 0) if inventory_item else 0

        return render_template("edit_product.html", product=product)

    return render_template("register_product.html", message="❌ Producto no encontrado.")


# Eliminar productos
@app.route("/delete-product", methods=["POST"])
def delete_product():
    clave = request.form["clave"]

    # --- CAMBIO CLAVE: Eliminar producto e inventario de Firestore ---
    product_deleted = delete_product_from_firestore(clave)
    inventory_deleted = delete_inventory_item_from_firestore(clave)  # También eliminar del inventario

    if product_deleted:  # Y si el inventario también se eliminó correctamente
        return redirect("/dashboard?message=✅ Producto eliminado correctamente.")
    else:
        return "❌ Error al eliminar producto."


# --- FUNCIONES PARA INTERACTUAR CON FIRESTORE (Clientes y Ventas) ---

def get_client_from_firestore(clave):
    """
    Recupera un cliente de la colección 'clients' por su clave.
    Retorna el diccionario de datos del cliente o None si no existe.
    """
    try:
        client_ref = db.collection('clients').document(clave)
        client_doc = client_ref.get()
        if client_doc.exists:
            return client_doc.to_dict()
        else:
            return None
    except Exception as e:
        print(f"Error al obtener cliente '{clave}' de Firestore: {e}")
        return None

def add_client_to_firestore(clave, client_data):
    """
    Agrega un nuevo cliente a la colección 'clients'.
    Usa la clave del cliente como ID del documento.
    Retorna True si la operación fue exitosa, False en caso contrario.
    """
    try:
        db.collection('clients').document(clave).set(client_data)
        return True
    except Exception as e:
        print(f"Error al agregar cliente '{clave}' a Firestore: {e}")
        return False

def update_client_in_firestore(clave, client_data):
    """
    Actualiza los datos de un cliente existente en la colección 'clients'.
    Retorna True si la operación fue exitosa, False en caso contrario.
    """
    try:
        db.collection('clients').document(clave).set(client_data) # set() with doc ID overwrites/updates
        return True
    except Exception as e:
        print(f"Error al actualizar cliente '{clave}' en Firestore: {e}")
        return False

def get_sales_by_client_from_firestore(clave_cliente):
    """
    Recupera todas las ventas para un cliente específico de la colección 'ventas'.
    Retorna una lista de diccionarios de ventas.
    """
    try:
        # Consulta todas las ventas donde 'cliente_clave' coincide
        sales = db.collection('ventas').where('cliente_clave', '==', clave_cliente).stream()
        sales_list = []
        for sale in sales:
            sales_list.append(sale.to_dict())
        return sales_list
    except Exception as e:
        print(f"Error al obtener ventas para el cliente '{clave_cliente}': {e}")
        return []

# --- FIN FUNCIONES PARA FIRESTORE (Clientes y Ventas) ---


# CLIENTES

# Registro de clientes
@app.route("/register-client", methods=["GET", "POST"])
def register_client():
    message = None

    if request.method == "POST":
        clave = request.form["clave"] # Ahora la clave viene del formulario

        # --- CAMBIO CLAVE: Verificar si el cliente ya existe en Firestore ---
        if get_client_from_firestore(clave):
            message = "❌ Error: Un cliente con esta clave ya existe."
        else:
            new_client = {
                "clave": clave,
                "nombre": request.form["nombre"],
                "apellido": request.form["apellido"],
                "direccion": request.form["direccion"],
                "telefono": request.form["telefono"],
                "correo": request.form["correo"],
                "credito_autorizado": request.form["credito_autorizado"] # Guarda como string, considera float
            }

            # --- CAMBIO CLAVE: Guardar nuevo cliente en Firestore ---
            if add_client_to_firestore(clave, new_client):
                message = "✅ Cliente registrado correctamente."
            else:
                message = "❌ Error al registrar el cliente en la base de datos."

    return render_template("registro_clientes.html", message=message)


# Busqueda de clientes
@app.route("/search-client", methods=["POST"])
def search_client():
    clave_busqueda = request.form["search_clave"]

    # --- CAMBIO CLAVE: Buscar el cliente en Firestore ---
    client = get_client_from_firestore(clave_busqueda)
    saldo_actual_cliente = None

    if client:
        # --- CAMBIO CLAVE: Calcular saldo del cliente usando Firebase ---
        saldo_actual_cliente = calcular_saldo_cliente(client["clave"])
        return render_template("edit_client.html", client=client, saldo_actual=saldo_actual_cliente, message="✅ Cliente encontrado.")
    else:
        return render_template("registro_clientes.html", message="❌ Cliente no encontrado.")


# Calculo de saldos (FUNCIÓN ADAPTADA PARA FIREBASE)
# Esta función es utilizada en el módulo de ventas y cobranza.
# Aseguraremos que siempre calcule el saldo DISPONIBLE del crédito.
def calcular_saldo_cliente(clave_cliente):
    # --- CAMBIO CLAVE: Obtener cliente de Firestore ---
    cliente = get_client_from_firestore(clave_cliente)
    credito_autorizado = 0.0
    saldo_pendiente_total = 0.0 # Este es el total de la DEUDA activa

    if cliente:
        try:
            # Asegurarse de que credito_autorizado sea un número
            credito_autorizado = float(cliente.get("credito_autorizado", 0))
        except ValueError:
            credito_autorizado = 0.0

        # --- ¡ESTE ES EL CAMBIO CRÍTICO! ---
        # Antes llamaba a get_sales_by_client_from_firestore (que trae TODAS las ventas).
        # Ahora debe llamar a get_pending_sales_for_client_from_firestore (que trae SOLO las ventas con saldo > 0).
        ventas_cliente = get_pending_sales_for_client_from_firestore(clave_cliente)

        for venta in ventas_cliente:
            try:
                # Sumar solo el 'saldo_a_pagar' de cada venta
                # Ya sabemos que estas ventas tienen saldo_a_pagar > 0 por la función de arriba.
                saldo_pendiente_total += float(venta.get("saldo_a_pagar", 0.0))
            except ValueError:
                pass

    # El saldo actual disponible es el crédito autorizado menos la deuda pendiente total
    saldo_actual = credito_autorizado - saldo_pendiente_total
    return saldo_actual # Retornar el saldo DISPONIBLE



# Encontrar Clientes (usado por AJAX para buscar un cliente individualmente)
@app.route("/get-client", methods=["GET"])
def get_client():
    clave = request.args.get("clave")

    # --- CAMBIO CLAVE: Buscar el cliente en Firestore ---
    client = get_client_from_firestore(clave)

    if client:
        # --- CAMBIO CLAVE: Calcular el saldo actual con la función Firebase-ready ---
        saldo_actual = calcular_saldo_cliente(clave)
        return jsonify({
            "found": True,
            "nombre": client["nombre"],
            "apellido": client["apellido"],
            "direccion": client["direccion"],
            "telefono": client["telefono"],
            "correo": client["correo"],
            "credito_autorizado": client["credito_autorizado"],
            "saldo_actual": saldo_actual
        })

    return jsonify({"found": False})


# Pagina de editar clientes
@app.route("/edit-client", methods=["POST", "GET"])
def edit_client():
    clave = request.form.get("clave") if request.method == "POST" else request.args.get("clave")

    # --- CAMBIO CLAVE: Obtener cliente de Firestore ---
    client = get_client_from_firestore(clave)
    saldo_actual_cliente = None

    if request.method == "POST" and client:
        print("Datos recibidos:", request.form)  # 🔍 Depuración para ver los datos enviados

        # Actualizar los campos del cliente
        client["nombre"] = request.form["nombre"]
        client["apellido"] = request.form["apellido"]
        client["direccion"] = request.form["direccion"]
        client["telefono"] = request.form["telefono"]
        client["correo"] = request.form["correo"]
        client["credito_autorizado"] = request.form["credito_autorizado"]

        # --- CAMBIO CLAVE: Guardar cambios en Firestore ---
        if update_client_in_firestore(clave, client):
            # Recalcular saldo para mostrar el más reciente
            saldo_actual_cliente = calcular_saldo_cliente(clave)
            return render_template("edit_client.html", client=client, saldo_actual=saldo_actual_cliente, message="✅ Cambios guardados correctamente.")
        else:
            saldo_actual_cliente = calcular_saldo_cliente(clave) # Recalcular incluso si falla para no perder el dato
            return render_template("edit_client.html", client=client, saldo_actual=saldo_actual_cliente, message="❌ Error al guardar los cambios en la base de datos.")

    if client:
        # --- CAMBIO CLAVE: Calcular saldo para la vista GET ---
        saldo_actual_cliente = calcular_saldo_cliente(clave)
        return render_template("edit_client.html", client=client, saldo_actual=saldo_actual_cliente)

    return render_template("registro_clientes.html", message="❌ Cliente no encontrado.")


# --- FUNCIONES PARA INTERACTUAR CON FIRESTORE (Inventario y Movimientos) ---

# Reutilizamos y ajustamos las funciones de productos si es necesario, pero las de inventario son clave
# Las funciones 'get_inventory_item_from_firestore' y 'update_inventory_item_in_firestore'
# ya las definimos en la sección de productos, por lo que las mantendremos.

# Función adicional para registrar movimientos (entradas, salidas, devoluciones)
def add_inventory_movement(collection_name, movement_data):
    """
    Agrega un nuevo movimiento de inventario (entrada, salida, devolución)
    a la colección especificada.
    Retorna True si la operación fue exitosa, False en caso contrario.
    """
    try:
        # Firestore generará un ID automático para cada movimiento
        db.collection(collection_name).add(movement_data)
        return True
    except Exception as e:
        print(f"Error al agregar movimiento a '{collection_name}': {e}")
        return False

# --- FIN FUNCIONES PARA FIRESTORE (Inventario y Movimientos) ---


# INVENTARIOS

# Pagina de entradas
@app.route("/inventory-entry", methods=["GET", "POST"])
def inventory_entry():
    message = None

    if request.method == "POST":
        clave = request.form["clave"]
        cantidad = float(request.form["cantidad"])
        costo_unitario = float(request.form["costo_unitario"])
        precio_unitario = float(request.form["precio_unitario"])
        fecha_hora_entrada = datetime.now().isoformat()

        # --- CAMBIO CLAVE: Obtener ítem de inventario de Firestore ---
        inventory_item = get_inventory_item_from_firestore(clave)

        if inventory_item:
            # Producto existente en inventario
            current_existencias = inventory_item.get("existencias", 0)
            new_existencias = current_existencias + cantidad

            # Actualizar datos en el documento de inventario
            inventory_item["existencias"] = new_existencias
            inventory_item["costo_unitario"] = costo_unitario  # Actualiza costo
            inventory_item["precio_unitario"] = precio_unitario  # Actualiza precio

            # --- CAMBIO CLAVE: Actualizar el ítem de inventario en Firestore ---
            success_update = update_inventory_item_in_firestore(clave, inventory_item)

            # Registrar el movimiento de entrada en la colección 'inventory_entries'
            entry_data = {
                "clave_producto": clave,  # Usamos clave_producto para que sea fácil consultar
                "fecha_hora": fecha_hora_entrada,
                "cantidad": cantidad,
                "costo_unitario": costo_unitario,
                "precio_unitario": precio_unitario
            }
            success_movement = add_inventory_movement("inventory_entries", entry_data)

            if success_update and success_movement:
                message = f"✅ Entrada registrada el {datetime.fromisoformat(fecha_hora_entrada).strftime('%Y-%m-%d %H:%M:%S')}. Existencias actuales: {new_existencias}."
            else:
                message = "❌ Error al registrar la entrada o actualizar inventario."

        else:
            # Si el producto no está en la colección 'inventory' (primera entrada)
            # Asegúrate de que el producto exista en la colección 'products' antes de añadirlo al inventario
            product_exists = get_product_from_firestore(clave)  # Reutilizamos la función de productos
            if not product_exists:
                message = f"❌ Error: El producto con clave '{clave}' no existe en la base de productos."
            else:
                new_inventory_item = {
                    "clave": clave,
                    "existencias": cantidad,
                    "costo_unitario": costo_unitario,
                    "precio_unitario": precio_unitario
                }
                # --- CAMBIO CLAVE: Agregar el nuevo ítem de inventario a Firestore ---
                success_add = update_inventory_item_in_firestore(clave,
                                                                 new_inventory_item)  # update_inventory_item_in_firestore también crea si no existe

                # Registrar el movimiento de entrada en la colección 'inventory_entries'
                entry_data = {
                    "clave_producto": clave,
                    "fecha_hora": fecha_hora_entrada,
                    "cantidad": cantidad,
                    "costo_unitario": costo_unitario,
                    "precio_unitario": precio_unitario
                }
                success_movement = add_inventory_movement("inventory_entries", entry_data)

                if success_add and success_movement:
                    message = f"✅ Entrada registrada el {datetime.fromisoformat(fecha_hora_entrada).strftime('%Y-%m-%d %H:%M:%S')}. Existencias actuales: {cantidad}."
                else:
                    message = "❌ Error al registrar la entrada o inicializar inventario."

    return render_template("inventory_entry.html", message=message)


# Pagina de Salidas y Devoluciones (No necesita cambios en esta ruta en sí)
@app.route("/inventory-management")
def inventory_management_page():
    return render_template("inventory_management.html")


# Salidas
@app.route("/inventory-exit", methods=["POST"])
def inventory_exit():
    message = None

    if request.method == "POST":
        clave = request.form["clave_salida"]
        cantidad_salida = float(request.form["cantidad_salida"])
        motivo_salida = request.form.get("motivo_salida", "Sin motivo")
        fecha_hora_salida = datetime.now().isoformat()

        # --- CAMBIO CLAVE: Obtener ítem de inventario de Firestore ---
        inventory_item = get_inventory_item_from_firestore(clave)

        if inventory_item:
            current_existencias = inventory_item.get("existencias", 0)
            if current_existencias >= cantidad_salida:
                new_existencias = current_existencias - cantidad_salida
                inventory_item["existencias"] = new_existencias

                # --- CAMBIO CLAVE: Actualizar el ítem de inventario en Firestore ---
                success_update = update_inventory_item_in_firestore(clave, inventory_item)

                # Registrar el movimiento de salida en la colección 'inventory_exits'
                exit_data = {
                    "clave_producto": clave,
                    "fecha_hora": fecha_hora_salida,
                    "cantidad": cantidad_salida,
                    "motivo": motivo_salida
                }
                success_movement = add_inventory_movement("inventory_exits", exit_data)

                if success_update and success_movement:
                    message = f"📤 Salida de {cantidad_salida} unidades de '{clave}' registrada el {datetime.fromisoformat(fecha_hora_salida).strftime('%Y-%m-%d %H:%M:%S')}. Existencias actuales: {new_existencias}."
                else:
                    message = "❌ Error al registrar la salida o actualizar inventario."
            else:
                message = f"⚠️ No hay suficientes existencias de '{clave}' para realizar la salida (disponibles: {current_existencias})."
        else:
            message = f"❌ No se encontró el producto con clave '{clave}' en el inventario."

    return render_template("inventory_management.html", message=message)


# Devoluciones
@app.route("/inventory-return", methods=["POST"])
def inventory_return():
    message = None

    if request.method == "POST":
        clave_devolucion = request.form["clave_devolucion"]
        cantidad_devolucion = float(request.form["cantidad_devolucion"])
        motivo_devolucion = request.form.get("motivo_devolucion", "Sin motivo")
        fecha_hora_devolucion = datetime.now().isoformat()

        # --- CAMBIO CLAVE: Obtener ítem de inventario de Firestore ---
        inventory_item = get_inventory_item_from_firestore(clave_devolucion)

        if inventory_item:
            current_existencias = inventory_item.get("existencias", 0)
            new_existencias = current_existencias + cantidad_devolucion
            inventory_item["existencias"] = new_existencias

            # --- CAMBIO CLAVE: Actualizar el ítem de inventario en Firestore ---
            success_update = update_inventory_item_in_firestore(clave_devolucion, inventory_item)

            # Registrar el movimiento de devolución en la colección 'inventory_returns'
            return_data = {
                "clave_producto": clave_devolucion,
                "fecha_hora": fecha_hora_devolucion,
                "cantidad": cantidad_devolucion,
                "motivo": motivo_devolucion
            }
            success_movement = add_inventory_movement("inventory_returns", return_data)

            if success_update and success_movement:
                message = f"🔄 Devolución de {cantidad_devolucion} unidades de '{clave_devolucion}' registrada el {datetime.fromisoformat(fecha_hora_devolucion).strftime('%Y-%m-%d %H:%M:%S')}. Existencias actuales: {new_existencias}."
            else:
                message = "❌ Error al registrar la devolución o actualizar inventario."
        else:
            message = f"❌ No se encontró el producto con clave '{clave_devolucion}' en el inventario."

    return render_template("inventory_management.html", message=message)


# --- FUNCIONES ADICIONALES PARA INTERACTUAR CON FIRESTORE (Ventas) ---

def get_sale_from_firestore(sale_id):
    """
    Recupera una venta de la colección 'sales' por su ID (folio).
    Retorna el diccionario de datos de la venta o None si no existe.
    """
    try:
        # Assuming sale_id is the document ID in Firestore for a sale
        sale_doc = db.collection('sales').document(sale_id).get()
        if sale_doc.exists:
            return sale_doc.to_dict()
        else:
            return None
    except Exception as e:
        print(f"Error al obtener venta '{sale_id}' de Firestore: {e}")
        return None

def add_sale_to_firestore(sale_data, custom_id=None):
    """
    Agrega una nueva venta a la colección 'sales'.
    Si custom_id es proporcionado, lo usa como ID del documento; de lo contrario,
    Firestore genera uno automáticamente.
    Retorna True y el ID de la venta si fue exitosa, False y None en caso contrario.
    """
    try:
        if custom_id:
            db.collection('sales').document(custom_id).set(sale_data)
            return True, custom_id
        else:
            # Let Firestore generate an ID
            doc_ref = db.collection('sales').add(sale_data)
            return True, doc_ref[1].id # doc_ref[1] is the DocumentReference, .id gets its ID
    except Exception as e:
        print(f"Error al agregar venta a Firestore: {e}")
        return False, None

def decrement_inventory_in_firestore(clave_producto, cantidad_vendida):
    """
    Decrementa las existencias de un producto en la colección 'inventory' y registra una salida.
    Utiliza una transacción para asegurar la atomicidad.
    Retorna True si la operación fue exitosa, False en caso contrario.
    """
    # Usamos una transacción para asegurar que la lectura y escritura de existencias sean atómicas
    transaction = db.transaction()
    inventory_ref = db.collection('inventory').document(clave_producto)

    @firestore.transactional
    def update_and_log_exit(transaction, inventory_ref, cantidad_vendida, clave_producto):
        snapshot = inventory_ref.get(transaction=transaction)
        if not snapshot.exists:
            raise ValueError(f"Producto '{clave_producto}' no encontrado en inventario para actualizar.")

        current_existencias = snapshot.get("existencias")
        if current_existencias is None:
            current_existencias = 0 # Default if field is missing

        if current_existencias < cantidad_vendida:
            raise ValueError(f"Existencias insuficientes para '{clave_producto}'. Disponible: {current_existencias}, Solicitado: {cantidad_vendida}.")

        new_existencias = current_existencias - cantidad_vendida
        transaction.update(inventory_ref, {"existencias": new_existencias})

        # Registrar la salida en la colección 'inventory_exits'
        exit_data = {
            "clave_producto": clave_producto,
            "fecha_hora": datetime.now().isoformat(),
            "cantidad": cantidad_vendida,
            "motivo": "Venta" # Motivo por defecto
        }
        db.collection('inventory_exits').add(exit_data) # Add directly, outside of transaction scope but after update

    try:
        update_and_log_exit(transaction, inventory_ref, cantidad_vendida, clave_producto)
        return True
    except ValueError as ve:
        print(f"Error de validación al decrementar inventario: {ve}")
        return False
    except Exception as e:
        print(f"Error inesperado al decrementar inventario para '{clave_producto}': {e}")
        return False

# --- FIN FUNCIONES ADICIONALES PARA FIRESTORE (Ventas) ---


# VENTAS

# Página de ventas
@app.route("/sales", methods=["GET", "POST"])
def sales():
    message = None

    # --- CAMBIO CLAVE: Cargar clientes y productos de Firestore ---
    # get_data_from_firestore('clients') buscaría todos los documentos en la colección 'clients'
    # db.collection('clients').stream() para obtener todos
    clients = [doc.to_dict() for doc in db.collection('clients').stream()]
    products = [doc.to_dict() for doc in db.collection('products').stream()]  # Cargar todos los productos disponibles

    if request.method == "POST":
        clave_venta = request.form["venta_id"]

        # --- CAMBIO CLAVE: Buscar la venta existente en Firestore ---
        venta_existente = get_sale_from_firestore(clave_venta)

        if venta_existente:
            cliente_clave_venta = venta_existente.get("cliente_clave")
            cliente = next((c for c in clients if c["clave"] == cliente_clave_venta), None)

            productos_en_venta = []
            for item in venta_existente.get("productos", []):
                # Aunque ya tienes descripción en la venta, podrías cargarla de products si quieres la última info
                producto_info = next((p for p in products if p["clave"] == item["clave"]), None)
                descripcion = producto_info["articulo"] if producto_info else "Artículo Desconocido"

                productos_en_venta.append({
                    "clave": item["clave"],
                    "descripcion": descripcion,  # Usa la descripción del producto o la guardada en la venta
                    "cantidad": item["cantidad"],
                    "precio_unitario": item["precio_unitario"],
                    "descuento": item.get("descuento", 0),
                    "impuesto": item.get("impuesto", 0),
                    "total": item.get("total", 0)  # Asegúrate de que el total por línea de producto esté guardado
                })

            return render_template("sales.html", client=cliente, products_in_sale=productos_en_venta,
                                   message="✅ Venta encontrada.", clients=clients,
                                   products=products)  # Pasar todas las listas para recargar el formulario
        else:
            message = "❌ Venta no encontrada."

    return render_template("sales.html", clients=clients, products=products, message=message)


# Procesar venta
@app.route("/process-sale", methods=["POST"])
def process_sale():
    message = None

    cliente_clave = request.form.get("cliente_clave")
    metodo_pago = request.form.get("metodo_pago")
    total_venta = float(request.form.get("total", "0"))  # Total final de la venta
    notas = request.form.get("notas", "").strip()

    # --- Obtener cliente de Firestore ---
    cliente = get_client_from_firestore(cliente_clave)
    if not cliente:
        # Re-fetch clients and products for the template to avoid errors
        clients_for_template = [doc.to_dict() for doc in db.collection('clients').stream()]
        products_for_template = [doc.to_dict() for doc in db.collection('products').stream()]
        return render_template("sales.html", message="❌ Error: Cliente no encontrado.",
                               clients=clients_for_template,
                               products=products_for_template)

    # Capturar la cantidad pagada inicialmente y el saldo a pagar inicial.
    # Es crucial que estos campos existan y se pasen desde el formulario para la CREACIÓN de la venta.
    try:
        # Asegurarse de que estos campos vienen del formulario de venta
        initial_cantidad_pagada = float(request.form.get("cantidad_pagada", "0"))
        initial_saldo_a_pagar = float(request.form.get("saldo_a_pagar", "0"))
    except ValueError:
        clients_for_template = [doc.to_dict() for doc in db.collection('clients').stream()]
        products_for_template = [doc.to_dict() for doc in db.collection('products').stream()]
        return render_template("sales.html", message="❌ Error: Los valores numéricos de pago/saldo no son válidos.",
                               clients=clients_for_template,
                               products=products_for_template)

    # Verificar el saldo de crédito si el método de pago es 'credito'
    if metodo_pago.lower() == 'credito':
        saldo_disponible = calcular_saldo_cliente(cliente_clave)
        if saldo_disponible < total_venta:
            clients_for_template = [doc.to_dict() for doc in db.collection('clients').stream()]
            products_for_template = [doc.to_dict() for doc in db.collection('products').stream()]
            return render_template("sales.html",
                                   message=f"❌ Error: Saldo de crédito insuficiente. Saldo disponible: {saldo_disponible:.2f}, Total de la venta: {total_venta:.2f}",
                                   clients=clients_for_template,
                                   products=products_for_template)

    productos_para_venta = []  # Detalles de los productos que se guardarán en la venta
    productos_a_decrementar_inventario = []  # Clave y cantidad para actualizar inventario

    form_productos_claves = request.form.getlist("clave[]")
    form_cantidades_str = request.form.getlist("cantidad[]")
    form_precios = request.form.getlist("precio[]")
    form_descuentos = request.form.getlist("descuento[]")
    form_impuestos = request.form.getlist("impuesto[]")

    fecha_hora_venta = datetime.now().isoformat()

    # Preparar detalles de la venta
    nombre_cliente = cliente.get("nombre", "")
    apellido_cliente = cliente.get("apellido", "")

    for i in range(len(form_productos_claves)):
        clave_producto = form_productos_claves[i]
        try:
            cantidad_solicitada = float(form_cantidades_str[i])
            precio_unitario = float(form_precios[i])
            descuento = float(form_descuentos[i])
            impuesto = float(form_impuestos[i])
        except ValueError:
            clients_for_template = [doc.to_dict() for doc in db.collection('clients').stream()]
            products_for_template = [doc.to_dict() for doc in db.collection('products').stream()]
            return render_template("sales.html",
                                   message=f"❌ Error: Datos numéricos inválidos para el producto {clave_producto}.",
                                   clients=clients_for_template,
                                   products=products_for_template)

        # --- Obtener producto e inventario de Firestore ---
        product_info = get_product_from_firestore(clave_producto)
        inventory_item = get_inventory_item_from_firestore(clave_producto)

        if not product_info or not inventory_item:
            clients_for_template = [doc.to_dict() for doc in db.collection('clients').stream()]
            products_for_template = [doc.to_dict() for doc in db.collection('products').stream()]
            return render_template("sales.html",
                                   message=f"❌ Error: Producto {clave_producto} no encontrado en base de datos o inventario.",
                                   clients=clients_for_template,
                                   products=products_for_template)

        existencias_producto = inventory_item.get("existencias", 0.0)

        if cantidad_solicitada > existencias_producto:
            clients_for_template = [doc.to_dict() for doc in db.collection('clients').stream()]
            products_for_template = [doc.to_dict() for doc in db.collection('products').stream()]
            return render_template("sales.html",
                                   message=f"❌ Error: No hay suficientes existencias para el producto {product_info.get('articulo', '')} ({clave_producto}). Solicitado: {cantidad_solicitada}, Disponible: {existencias_producto}",
                                   clients=clients_for_template,
                                   products=products_for_template)

        # Calcular total por línea de producto
        total_linea = (cantidad_solicitada * precio_unitario) * (1 - descuento / 100) * (
                    1 + impuesto / 100)  # Asumiendo descuento e impuesto como porcentajes

        productos_para_venta.append({
            "clave": clave_producto,
            "nombre_articulo": product_info.get("articulo", ""),
            "cantidad": cantidad_solicitada,
            "precio_unitario": precio_unitario,
            "descuento": descuento,
            "impuesto": impuesto,
            "total_linea": total_linea
        })
        productos_a_decrementar_inventario.append({"clave": clave_producto, "cantidad": cantidad_solicitada})

    # Construir el documento de la nueva venta
    nueva_venta_data = {
        "fecha_hora": fecha_hora_venta,
        "cliente_clave": cliente_clave,
        "nombre_cliente": nombre_cliente,
        "apellido_cliente": apellido_cliente,
        "productos": productos_para_venta,
        "total": total_venta,
        "metodo_pago": metodo_pago,
        "cantidad_pagada": initial_cantidad_pagada,  # <-- Usa la cantidad pagada del formulario
        "saldo_a_pagar": initial_saldo_a_pagar,  # <-- Usa el saldo a pagar del formulario
        "notas": notas,
        "pagos_realizados": []  # Inicializa la lista de pagos para esta venta
    }

    # Si la venta es a crédito y el saldo a pagar es el total de la venta,
    # significa que no se pagó nada inicialmente (o se pagó 0).
    # Ajusta 'cantidad_pagada' si es una venta a crédito completa y no se ingresó pago inicial.
    if metodo_pago.lower() == 'credito' and initial_cantidad_pagada == 0 and initial_saldo_a_pagar == total_venta:
        pass  # No hacer nada, los valores iniciales son correctos.
    elif initial_cantidad_pagada > 0:
        # Si hay un pago inicial, regístralo como el primer pago.
        nueva_venta_data["pagos_realizados"].append({
            "fecha_hora": fecha_hora_venta,  # Mismo timestamp que la venta
            "monto": initial_cantidad_pagada
        })

    # --- Guardar la venta en Firestore y obtener su ID (folio) ---
    # Usaremos el ID generado por Firestore como el 'folio' si no pasas un custom_id.
    # Esto es más robusto que un contador simple si hay muchas operaciones concurrentes.
    success_sale_add, sale_firestore_id = add_sale_to_firestore(nueva_venta_data)

    if not success_sale_add:
        clients_for_template = [doc.to_dict() for doc in db.collection('clients').stream()]
        products_for_template = [doc.to_dict() for doc in db.collection('products').stream()]
        message = "❌ Error: No se pudo registrar la venta en la base de datos."
        return render_template("sales.html", message=message,
                               clients=clients_for_template,
                               products=products_for_template)

    # Después de guardar la venta, actualiza el inventario para cada producto vendido
    inventory_update_success = True
    for item_vendido in productos_a_decrementar_inventario:
        if not decrement_inventory_in_firestore(item_vendido["clave"], item_vendido["cantidad"]):
            inventory_update_success = False
            message = f"❌ Error: Falló la actualización de inventario para el producto {item_vendido['clave']}."
            # Considerar aquí una estrategia de compensación/reversión de la venta si el inventario es crítico.
            break

    if inventory_update_success:
        message = f"✅ Venta procesada exitosamente. Folio: {sale_firestore_id}"
    else:
        message = f"⚠️ Venta registrada con folio {sale_firestore_id}, pero hubo un error al actualizar el inventario de algunos productos. ¡Revisa!"

    clients_for_template = [doc.to_dict() for doc in db.collection('clients').stream()]
    products_for_template = [doc.to_dict() for doc in db.collection('products').stream()]
    return render_template("sales.html", message=message,
                           clients=clients_for_template,
                           products=products_for_template)


# Ver informacion ventas (GET)
@app.route("/get-sale", methods=["GET"])
def get_sale():
    clave_venta = request.args.get("clave")  # Asumiendo que es el ID del documento de venta en Firestore

    # --- CAMBIO CLAVE: Obtener venta de Firestore ---
    venta = get_sale_from_firestore(clave_venta)

    if venta:
        # --- CAMBIO CLAVE: Cargar clientes y productos de Firestore para obtener nombres/descripciones ---
        clients_data = {c.get("clave"): c for c in [doc.to_dict() for doc in db.collection('clients').stream()]}
        products_data = {p.get("clave"): p for p in [doc.to_dict() for doc in db.collection('products').stream()]}
        inventory_data = {item.get("clave"): item for item in
                          [doc.to_dict() for doc in db.collection('inventory').stream()]}

        cliente_info = clients_data.get(venta.get("cliente_clave"), {})

        productos_en_venta = []
        for item in venta.get("productos", []):
            producto_detail = products_data.get(item["clave"], {})
            inventory_detail = inventory_data.get(item["clave"], {})

            productos_en_venta.append({
                "clave": item["clave"],
                "descripcion": producto_detail.get("articulo", "N/A"),  # Usar 'articulo' del producto
                "existencias": inventory_detail.get("existencias", 0),  # Existencias actuales
                "cantidad": item["cantidad"],
                "precio_unitario": item["precio_unitario"],
                "descuento": item.get("descuento", 0),
                "impuesto": item.get("impuesto", 0),
                "total_linea": item.get("total_linea", 0)  # Si guardaste este campo
            })

        return jsonify({
            "found": True,
            "folio": clave_venta,  # O venta.get("folio") si lo guardaste diferente
            "fecha_hora": venta.get("fecha_hora"),
            "cliente_clave": venta.get("cliente_clave"),
            "nombre_cliente": cliente_info.get("nombre", "") + " " + cliente_info.get("apellido", ""),
            "total_venta": venta.get("total"),
            "metodo_pago": venta.get("metodo_pago"),
            "cantidad_pagada": venta.get("cantidad_pagada"),
            "saldo_a_pagar": venta.get("saldo_a_pagar"),
            "notas": venta.get("notas", ""),
            "productos": productos_en_venta
        })

    return jsonify({"found": False})


# Actualizar existencias (Esta función ya no es necesaria como separada, su lógica está en decrement_inventory_in_firestore)
# def update_inventory(clave_producto, cantidad_vendida):
#     # Esta lógica ahora está encapsulada en decrement_inventory_in_firestore y usa transacciones.
#     pass


# --- FUNCIONES ADICIONALES PARA INTERACTUAR CON FIRESTORE (Cobranza) ---

# get_client_from_firestore(clave) - Ya definida en el módulo de clientes
# get_sale_from_firestore(sale_id) - Ya definida en el módulo de ventas

def get_pending_sales_for_client_from_firestore(clave_cliente):
    """
    Recupera las ventas pendientes de pago para un cliente específico.
    Retorna una lista de diccionarios de ventas.
    """
    try:
        # Consulta las ventas para el cliente donde 'saldo_a_pagar' es mayor que 0
        sales = db.collection('sales') \
            .where('cliente_clave', '==', clave_cliente) \
            .where('saldo_a_pagar', '>', 0) \
            .stream()

        pending_sales_list = []
        for sale in sales:
            sale_data = sale.to_dict()
            sale_data['id'] = sale.id  # Incluir el ID del documento de Firestore como 'folio'
            pending_sales_list.append(sale_data)
        return pending_sales_list
    except Exception as e:
        print(f"Error al obtener ventas pendientes para el cliente '{clave_cliente}': {e}")
        return []


def update_sale_payment_in_firestore(sale_id, updated_sale_data):
    """
    Actualiza el documento de una venta en Firestore con los nuevos datos de pago.
    Retorna True si la operación fue exitosa, False en caso contrario.
    """
    try:
        db.collection('sales').document(sale_id).set(updated_sale_data)  # set() sobrescribe/actualiza
        return True
    except Exception as e:
        print(f"Error al actualizar pago para la venta '{sale_id}' en Firestore: {e}")
        return False


# --- FIN FUNCIONES PARA FIRESTORE (Cobranza) ---


# COBRANZA

# Pagina de Cobranza (No requiere cambios directos aquí)
@app.route("/billing")
def billing():
    if 'user' in session:
        return render_template("cobranza.html")
    # Puedes añadir un redirect si no hay usuario en sesión
    return redirect("/")



# Devolver datos del cliente y sus deudas
@app.route("/get-client-debts", methods=["GET"])
def get_client_debts():
    clave_cliente = request.args.get("clave")
    print(f"DEBUG_COBRANZA: Buscando deudas para cliente: {clave_cliente}")

    cliente = get_client_from_firestore(clave_cliente)
    if not cliente:
        print(f"DEBUG_COBRANZA: Cliente {clave_cliente} no encontrado.")
        return {"found": False}

    ventas_pendientes_info = []
    saldo_pendiente_total = 0.0

    ventas_pendientes_firestore = get_pending_sales_for_client_from_firestore(clave_cliente)
    print(f"DEBUG_COBRANZA: Ventas pendientes encontradas para {clave_cliente}: {len(ventas_pendientes_firestore)}")

    for venta in ventas_pendientes_firestore:
        folio_venta = venta.get("folio", venta.get("id"))
        saldo_a_pagar = float(venta.get("saldo_a_pagar", 0.0))
        print(f"DEBUG_COBRANZA: Venta: {folio_venta}, saldo_a_pagar: {saldo_a_pagar}")

        if saldo_a_pagar > 0:
            ventas_pendientes_info.append({
                "folio": folio_venta,
                "saldo_a_pagar": saldo_a_pagar
            })
            saldo_pendiente_total += saldo_a_pagar

    print(f"DEBUG_COBRANZA: Saldo pendiente total calculado: {saldo_pendiente_total}")
    return jsonify({
        "found": True,
        "nombre": cliente.get("nombre", ""),
        "apellido": cliente.get("apellido", ""),
        "saldo_pendiente_total": saldo_pendiente_total,
        "ventas_pendientes": ventas_pendientes_info
    })


# Procesar pagos
@app.route("/process-payment", methods=["POST"])
def process_payment():
    data = request.get_json()
    cliente_clave = data.get("cliente_clave")
    pagos = data.get("pagos", {})  # Diccionario con {folio_venta: monto_pagado}

    if not cliente_clave or not pagos:
        return jsonify({"success": False, "message": "Error: Faltan datos para procesar el pago."})

    # --- CAMBIO CLAVE: Obtener todas las ventas pendientes del cliente ---
    ventas_pendientes_firestore = get_pending_sales_for_client_from_firestore(cliente_clave)

    # Crear un mapeo de folio a documento de venta para fácil acceso
    sales_map = {sale.get("folio", sale.get("id")): sale for sale in ventas_pendientes_firestore}

    successful_updates = 0
    total_payments = len(pagos)

    for folio_venta, cantidad_pagada_str in pagos.items():
        try:
            cantidad_pagada = float(cantidad_pagada_str)
        except ValueError:
            print(f"Advertencia: Monto de pago inválido para folio {folio_venta}: {cantidad_pagada_str}")
            continue  # Saltar a la siguiente venta

        current_sale_data = sales_map.get(folio_venta)

        if current_sale_data:
            fecha_hora_pago = datetime.now().isoformat()

            # Asegurarse de que exista una lista de pagos para esta venta
            if "pagos_realizados" not in current_sale_data:
                current_sale_data["pagos_realizados"] = []

            current_sale_data["pagos_realizados"].append({
                "fecha_hora": fecha_hora_pago,
                "monto": cantidad_pagada
            })

            # Asegurarse de que 'cantidad_pagada' y 'saldo_a_pagar' sean float antes de operar
            current_sale_data["cantidad_pagada"] = float(
                current_sale_data.get("cantidad_pagada", 0.0)) + cantidad_pagada
            current_sale_data["saldo_a_pagar"] = float(current_sale_data.get("saldo_a_pagar", 0.0)) - cantidad_pagada

            if current_sale_data["saldo_a_pagar"] < 0:
                current_sale_data["saldo_a_pagar"] = 0.0  # Evitar saldos negativos

            # --- CAMBIO CLAVE: Actualizar el documento de venta en Firestore ---
            sale_doc_id = current_sale_data.get("id")  # Usar el ID del documento de Firestore
            if update_sale_payment_in_firestore(sale_doc_id, current_sale_data):
                successful_updates += 1
            else:
                print(f"Error al actualizar el pago para la venta {folio_venta} en Firestore.")
        else:
            print(f"Advertencia: Venta con folio {folio_venta} no encontrada para el cliente {cliente_clave}.")

    if successful_updates == total_payments:
        return jsonify({"success": True,
                        "message": f"✅ Pagos procesados exitosamente para {successful_updates} de {total_payments} ventas."})
    elif successful_updates > 0:
        return jsonify({"success": True,
                        "message": f"⚠️ Pagos procesados parcialmente ({successful_updates} de {total_payments} ventas). Algunos pagos no se pudieron aplicar."})
    else:
        return jsonify({"success": False, "message": "❌ No se pudo procesar ningún pago. Verifica los datos."})


# --- FUNCIONES DE AYUDA PARA OBTENER TODOS LOS DATOS DE UNA COLECCIÓN ---

def get_all_documents_from_collection(collection_name):
    """
    Obtiene todos los documentos de una colección de Firestore.
    Retorna una lista de diccionarios, cada uno con los datos del documento
    y su 'id' de Firestore.
    """
    try:
        docs = db.collection(collection_name).stream()
        data = []
        for doc in docs:
            doc_data = doc.to_dict()
            doc_data['id'] = doc.id # Incluir el ID del documento, útil para ventas
            data.append(doc_data)
        return data
    except Exception as e:
        print(f"Error al obtener todos los documentos de '{collection_name}': {e}")
        return []

# --- FIN FUNCIONES DE AYUDA ---

# REPORTES

# Ir Pagina de Reportes (No requiere cambios)
@app.route('/reportes')
def reportes():
    if 'user' in session:
        return render_template('reportes.html')
    return redirect("/")


# Exportar Ventas con Notas
@app.route('/reportes/export/ventas')
def export_sales_excel():
    # --- CAMBIO CLAVE: Cargar ventas de Firestore ---
    ventas_data = get_all_documents_from_collection("sales")
    # --- CAMBIO CLAVE: Cargar productos para obtener nombres de artículos ---
    products_data = get_all_documents_from_collection("products")
    productos_dict = {p.get('clave'): p.get('articulo') for p in products_data if p.get('clave')}

    wb = Workbook()
    ws = wb.active
    ws.append(
        ['Folio', 'Fecha y Hora', 'Cliente Clave', 'Nombre Cliente', 'Artículos Vendidos', 'Total', 'Método de Pago',
         'Cantidad Pagada', 'Saldo Pendiente', 'Notas'])

    for venta in ventas_data:
        # Usar el ID del documento de Firestore como 'folio' si no se guardó un campo 'folio' explícito
        folio = venta.get('folio', venta.get('id', ''))

        articulos_vendidos = []
        for producto_en_venta in venta.get('productos', []):
            clave_prod = producto_en_venta.get('clave', '')
            nombre_articulo = productos_dict.get(clave_prod, producto_en_venta.get('nombre_articulo',
                                                                                   'No encontrado'))  # Preferir nombre guardado en venta, sino buscar por clave
            cantidad_vendida = producto_en_venta.get('cantidad', 0)
            articulos_vendidos.append(f"{nombre_articulo} (x{cantidad_vendida})")

        nombres_articulos = "\n".join(articulos_vendidos)
        notas = venta.get('notas', '')

        ws.append([
            folio,
            venta.get('fecha_hora', ''),
            venta.get('cliente_clave', ''),
            f"{venta.get('nombre_cliente', '')} {venta.get('apellido_cliente', '')}".strip(),
            nombres_articulos,
            venta.get('total', ''),
            venta.get('metodo_pago', ''),
            venta.get('cantidad_pagada', ''),
            venta.get('saldo_a_pagar', ''),
            notas
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, download_name='reporte_ventas.xlsx', as_attachment=True)


# Exportar Existencias
@app.route('/reportes/export/existencias')
def export_inventory_excel():
    # --- CAMBIO CLAVE: Cargar inventario de Firestore ---
    inventory_data = get_all_documents_from_collection("inventory")
    # --- CAMBIO CLAVE: Cargar productos para obtener nombres de artículos ---
    products_data = get_all_documents_from_collection("products")
    productos_dict = {p.get('clave'): p.get('articulo') for p in products_data if p.get('clave')}

    wb = Workbook()
    ws = wb.active
    ws.append(['Clave', 'Artículo', 'Existencias', 'Costo Unitario', 'Precio Unitario'])

    for item in inventory_data:
        clave_producto = item.get('clave', '')
        nombre_articulo = productos_dict.get(clave_producto, 'No encontrado')

        ws.append([
            clave_producto,
            nombre_articulo,
            item.get('existencias', ''),
            item.get('costo_unitario', ''),
            item.get('precio_unitario', '')
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, download_name='reporte_existencias.xlsx', as_attachment=True)


# Exportar Saldos de Clientes
@app.route('/reportes/export/saldos')
def export_balances_excel():
    # --- CAMBIO CLAVE: Cargar todas las ventas de Firestore ---
    ventas_data = get_all_documents_from_collection("sales")
    # --- CAMBIO CLAVE: Cargar todos los clientes de Firestore ---
    clientes_data = get_all_documents_from_collection("clients")
    clientes_dict = {c.get('clave'): {'nombre': c.get('nombre', ''), 'apellido': c.get('apellido', '')} for c in
                     clientes_data if c.get('clave')}

    wb = Workbook()
    ws = wb.active
    ws.append(['Cliente', 'Folio de Venta', 'Total Venta', 'Cantidad Pagada', 'Saldo Pendiente', 'Fecha y Hora Venta'])

    for venta in ventas_data:
        cliente_clave = venta.get('cliente_clave', '')
        cliente_info = clientes_dict.get(cliente_clave, {'nombre': 'No encontrado', 'apellido': ''})
        nombre_completo_cliente = f"{cliente_info['nombre']} {cliente_info.get('apellido', '')}".strip()

        ws.append([
            nombre_completo_cliente,
            venta.get('folio', venta.get('id', '')),  # Usar 'folio' si existe, sino el ID del documento
            venta.get('total', ''),
            venta.get('cantidad_pagada', ''),
            venta.get('saldo_a_pagar', ''),
            venta.get('fecha_hora', '')
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, download_name='reporte_saldos_clientes.xlsx', as_attachment=True)


# Exportar Catalogo de Productos
@app.route('/reportes/export/productos')
def export_products_excel():
    # --- CAMBIO CLAVE: Cargar productos de Firestore ---
    products_data = get_all_documents_from_collection("products")

    wb = Workbook()
    ws = wb.active
    ws.append(['Clave', 'Artículo', 'Marca', 'Categoría', 'Tamaño', 'Observaciones'])
    for product in products_data:
        ws.append([
            product.get('clave', ''),
            product.get('articulo', ''),
            product.get('marca', ''),
            product.get('categoria', ''),
            product.get('tamaño', ''),
            product.get('observaciones', '')
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, download_name='catalogo_productos.xlsx', as_attachment=True)


#  Exportar Base de Clientes
@app.route('/reportes/export/clientes')
def export_clients_excel():
    # --- CAMBIO CLAVE: Cargar clientes de Firestore ---
    clients_data = get_all_documents_from_collection("clients")

    wb = Workbook()
    ws = wb.active
    ws.append(['Clave', 'Nombre', 'Apellido', 'Dirección', 'Teléfono', 'Correo', 'Crédito Autorizado'])
    for client in clients_data:
        ws.append([
            client.get('clave', ''),
            client.get('nombre', ''),
            client.get('apellido', ''),
            client.get('direccion', ''),
            client.get('telefono', ''),
            client.get('correo', ''),
            client.get('credito_autorizado', '')
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, download_name='base_de_clientes.xlsx', as_attachment=True)


# Carga de ventas y genera gráfica de importes por día
def generate_sales_graph(start_date=None, end_date=None):
    # --- CAMBIO CLAVE: Cargar ventas de Firestore ---
    ventas_data = get_all_documents_from_collection("sales")

    ventas_filtradas = []
    if start_date and end_date:
        for venta in ventas_data:
            # Asegúrate de que 'fecha_hora' exista y sea un formato ISO
            fecha_venta_str = venta.get('fecha_hora', '').split('T')[0]
            if fecha_venta_str and start_date <= fecha_venta_str <= end_date:
                ventas_filtradas.append(venta)
        ventas_data = ventas_filtradas

    # Agrupar ventas por fecha y sumar los totales
    sales_by_date = {}
    for venta in ventas_data:
        fecha_venta_str = venta.get('fecha_hora', '').split('T')[0]
        try:
            total_venta = float(venta.get('total', 0))
            sales_by_date[fecha_venta_str] = sales_by_date.get(fecha_venta_str, 0) + total_venta
        except ValueError:
            print(f"Advertencia: Total de venta inválido para el folio {venta.get('folio', venta.get('id'))}")
            continue

    dates_sorted = sorted(sales_by_date.keys())
    total_sales = [sales_by_date[date] for date in dates_sorted]

    plt.figure(figsize=(10, 6))
    plt.plot(dates_sorted, total_sales, marker='o')
    plt.title('Importe Total de Ventas por Día')
    plt.xlabel('Fecha')
    plt.ylabel('Importe Total de Ventas')
    plt.grid(True)
    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()

    img = BytesIO()
    plt.savefig(img, format='png')
    img.seek(0)
    plot_url = base64.b64encode(img.getvalue()).decode('utf8')
    plt.close()

    return plot_url


# Generador de graficos de ventas (No requiere cambios en la ruta en sí, solo en la función que llama)
@app.route('/reportes/grafica/ventas')
def view_sales_graph():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    plot_url = generate_sales_graph(start_date, end_date)
    return render_template('graph_viewer.html', plot_url=plot_url,
                           report_title='Gráfica del Importe Total de Ventas por Día')


# Carga de inventarios (para gráfico)
def generate_inventory_graph():
    # --- CAMBIO CLAVE: Cargar inventario de Firestore ---
    inventory_data = get_all_documents_from_collection("inventory")
    # --- CAMBIO CLAVE: Cargar productos para obtener nombres de artículos ---
    products_data = get_all_documents_from_collection("products")
    productos_dict = {p.get('clave'): p.get('articulo') for p in products_data if p.get('clave')}

    product_names = []
    stock_levels = []
    for item in inventory_data:
        clave_producto = item.get('clave')
        nombre_articulo = productos_dict.get(clave_producto, 'Desconocido')
        product_names.append(nombre_articulo)
        try:
            stock_levels.append(float(item.get('existencias', 0)))
        except ValueError:
            stock_levels.append(0)  # Default to 0 if existencias is not a number

    plt.figure(figsize=(10, 6))
    plt.bar(product_names, stock_levels)
    plt.title('Niveles de Existencia por Producto')
    plt.xlabel('Producto')
    plt.ylabel('Cantidad en Stock')
    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()

    img = BytesIO()
    plt.savefig(img, format='png')
    img.seek(0)
    plot_url = base64.b64encode(img.getvalue()).decode('utf8')
    plt.close()

    return plot_url


# Grafico de Existencias (No requiere cambios en la ruta en sí, solo en la función que llama)
@app.route('/reportes/grafica/inventario')
def view_inventory_graph():
    plot_url = generate_inventory_graph()
    return render_template('graph_viewer.html', plot_url=plot_url, report_title='Gráfica de Niveles de Existencia')


# Carga de Saldos (para gráfico)
def generate_balances_graph():
    # --- CAMBIO CLAVE: Cargar todas las ventas de Firestore ---
    ventas_data = get_all_documents_from_collection("sales")
    # --- CAMBIO CLAVE: Cargar todos los clientes para obtener nombres ---
    clientes_data = get_all_documents_from_collection("clients")
    clientes_dict = {c.get('clave'): f"{c.get('nombre', '')} {c.get('apellido', '')}".strip() for c in clientes_data if
                     c.get('clave')}

    balances = {}
    for venta in ventas_data:
        cliente_clave = venta.get('cliente_clave')
        try:
            saldo = float(venta.get('saldo_a_pagar', 0))
            if cliente_clave:  # Asegurarse de que haya una clave de cliente
                balances[cliente_clave] = balances.get(cliente_clave, 0) + saldo
        except ValueError:
            print(f"Advertencia: Saldo a pagar inválido para el folio {venta.get('folio', venta.get('id'))}")
            continue

    clients_labels = [clientes_dict.get(c, 'Cliente Desconocido') for c in sorted(balances.keys())]
    saldo_pendiente = [balances[c] for c in sorted(balances.keys())]

    plt.figure(figsize=(10, 6))
    plt.bar(clients_labels, saldo_pendiente)
    plt.title('Saldos Pendientes por Cliente')
    plt.xlabel('Cliente')
    plt.ylabel('Saldo Pendiente')
    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()

    img = BytesIO()
    plt.savefig(img, format='png')
    img.seek(0)
    plot_url = base64.b64encode(img.getvalue()).decode('utf8')
    plt.close()

    return plot_url


# Grafico de Saldos (No requiere cambios en la ruta en sí, solo en la función que llama)
@app.route('/reportes/grafica/saldos')
def view_balances_graph():
    plot_url = generate_balances_graph()
    return render_template('graph_viewer.html', plot_url=plot_url,
                           report_title='Gráfica de Saldos Pendientes por Cliente')


# Ir a pagina de graficos (No requiere cambios)
@app.route('/graph_viewer')
def graph_viewer():
    plot_url = request.args.get('plot_url', '')
    report_title = request.args.get('report_title', 'Gráfica')
    return render_template('graph_viewer.html', plot_url=plot_url, report_title=report_title)




if __name__ == "__main__":
    if getattr(sys, 'frozen', False):
        threading.Thread(target=open_browser, daemon=True).start()
        app.run(debug=False, use_reloader=False)
    else:
        app.run(debug=True)